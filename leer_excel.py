"""Lectura y validación de parámetros desde un archivo Excel.

Este módulo encapsula la lógica necesaria para leer la hoja de cálculo
de entrada que contiene:
- Número de trabajadores Full-Time (celda C3)
- Número de trabajadores Part-Time (celda C4)
- Tipo de turno (A o B) en la celda C5
- Demanda semanal en las filas 8 a 14 (columnas C, D, E)

La función principal `leer_parametros` devuelve una tupla con los
valores leídos y un diccionario con la demanda por día.
"""

from openpyxl import load_workbook


def leer_parametros(path):
    """Lee el archivo Excel y retorna los parámetros y la demanda.

    Args:
        path (str): Ruta al archivo .xlsx con los parámetros.

    Returns:
        tuple: (full_time, part_time, turno, demanda)

    Raises:
        ValueError: Si las celdas C3/C4/C5 no contienen datos válidos.
    """
    # Abrir libro en modo data_only para obtener valores calculados
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Leer parámetros básicos (celdas definidas por convención)
    full_time = ws["C3"].value
    part_time = ws["C4"].value
    turno = str(ws["C5"].value).strip().upper()

    # Validación mínima. Se espera que turno sea 'A' o 'B'.
    if full_time is None or part_time is None or turno not in ["A", "B"]:
        raise ValueError("El Excel no contiene valores válidos en C3, C4 o C5.")

    # Construir diccionario de demanda por día leyendo filas consecutivas
    demanda = {}
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # Las filas donde se espera la demanda comienzan en la fila 8
    fila_inicio = 8
    for i, dia in enumerate(dias):
        fila = fila_inicio + i
        # Columnas C/D/E contienen Mañana/Intermedio/Tarde
        mañana = ws[f"C{fila}"].value
        inter = ws[f"D{fila}"].value
        tarde = ws[f"E{fila}"].value

        # Normalizar valores nulos a 0 para evitar errores posteriores
        demanda[dia] = {
            "Mañana": mañana if mañana else 0,
            "Intermedio": inter if inter else 0,
            "Tarde": tarde if tarde else 0,
        }

    return full_time, part_time, turno, demanda
