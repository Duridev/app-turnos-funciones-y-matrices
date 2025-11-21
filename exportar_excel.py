"""Módulo para exportar los horarios y turnos a Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def exportar_informe_completo(ruta_archivo, parametros, demanda, matriz_turnos, 
                               descanso_sab, descanso_dom, horarios_trabajadores):
    """
    Genera un archivo Excel con el informe completo de turnos.
    
    Args:
        ruta_archivo: Ruta donde guardar el archivo Excel
        parametros: Tupla (full_time, part_time, tipo)
        demanda: Diccionario con demanda por día
        matriz_turnos: Matriz 7x3 con asignación de turnos
        descanso_sab: Cantidad que descansa sábado
        descanso_dom: Cantidad que descansa domingo
        horarios_trabajadores: Diccionario con horarios por trabajador
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: Resumen General ===
    ws_resumen = wb.create_sheet("Resumen General")
    
    ws_resumen['A1'] = "INFORME DE TURNOS - CAFETERÍA"
    ws_resumen['A1'].font = titulo_font
    ws_resumen['A1'].fill = titulo_fill
    ws_resumen['A1'].alignment = center_align
    ws_resumen.merge_cells('A1:E1')
    
    ws_resumen['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumen.merge_cells('A2:E2')
    
    # Parámetros
    full_time, part_time, tipo = parametros
    ws_resumen['A4'] = "PARÁMETROS"
    ws_resumen['A4'].font = header_font
    ws_resumen['A4'].fill = header_fill
    ws_resumen.merge_cells('A4:B4')
    
    ws_resumen['A5'] = "Trabajadores Full-Time:"
    ws_resumen['B5'] = full_time
    ws_resumen['A6'] = "Trabajadores Part-Time:"
    ws_resumen['B6'] = part_time
    ws_resumen['A7'] = "Tipo de Turno:"
    ws_resumen['B7'] = tipo
    ws_resumen['B7'].alignment = Alignment(horizontal='right', vertical='center')
    ws_resumen['A8'] = "Descansan Sábado:"
    ws_resumen['B8'] = descanso_sab
    ws_resumen['A9'] = "Descansan Domingo:"
    ws_resumen['B9'] = descanso_dom
    
    # Demanda
    ws_resumen['D4'] = "DEMANDA SEMANAL"
    ws_resumen['D4'].font = header_font
    ws_resumen['D4'].fill = header_fill
    ws_resumen.merge_cells('D4:G4')
    
    ws_resumen['D5'] = "Día"
    ws_resumen['E5'] = "Mañana"
    ws_resumen['F5'] = "Intermedio"
    ws_resumen['G5'] = "Tarde"
    
    for col in ['D', 'E', 'F', 'G']:
        ws_resumen[f'{col}5'].font = Font(bold=True)
        ws_resumen[f'{col}5'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws_resumen[f'{col}5'].alignment = center_align
    
    fila = 6
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    for i, dia in enumerate(dias):
        ws_resumen[f'D{fila}'] = dia
        ws_resumen[f'E{fila}'] = int(matriz_turnos[i][0])
        ws_resumen[f'F{fila}'] = int(matriz_turnos[i][1])
        ws_resumen[f'G{fila}'] = int(matriz_turnos[i][2])
        fila += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 12
    ws_resumen.column_dimensions['F'].width = 12
    ws_resumen.column_dimensions['G'].width = 12
    
    # === HOJA 2: Horario por Trabajador ===
    ws_horarios = wb.create_sheet("Horario Semanal")
    
    ws_horarios['A1'] = "HORARIO SEMANAL POR TRABAJADOR"
    ws_horarios['A1'].font = titulo_font
    ws_horarios['A1'].fill = titulo_fill
    ws_horarios['A1'].alignment = center_align
    ws_horarios.merge_cells('A1:H1')
    
    # Encabezados
    ws_horarios['A3'] = "Trabajador"
    for i, dia in enumerate(dias):
        col = get_column_letter(i + 2)
        ws_horarios[f'{col}3'] = dia
    
    for col_idx in range(1, 9):
        col = get_column_letter(col_idx)
        ws_horarios[f'{col}3'].font = header_font
        ws_horarios[f'{col}3'].fill = header_fill
        ws_horarios[f'{col}3'].alignment = center_align
        ws_horarios[f'{col}3'].border = border
    
    # Datos - Separar Full-Time y Part-Time
    trabajadores_ft = []
    trabajadores_pt = []
    
    for trabajador in horarios_trabajadores.keys():
        if trabajador.startswith("Trabajador"):
            trabajadores_ft.append(trabajador)
        else:
            trabajadores_pt.append(trabajador)
    
    # Ordenar cada grupo
    trabajadores_ft.sort()
    trabajadores_pt.sort()
    
    # Combinar: primero Full-Time, luego Part-Time
    trabajadores_ordenados = trabajadores_ft + trabajadores_pt
    
    fila = 4
    for trabajador in trabajadores_ordenados:
        horario = horarios_trabajadores[trabajador]
        ws_horarios[f'A{fila}'] = trabajador
        ws_horarios[f'A{fila}'].border = border
        
        for i, dia in enumerate(dias):
            col = get_column_letter(i + 2)
            turno = horario[dia]
            celda = ws_horarios[f'{col}{fila}']
            celda.value = turno
            celda.alignment = center_align
            celda.border = border
            
            # Colorear según turno
            if turno == "Mañana":
                celda.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif turno == "Intermedio":
                celda.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            elif turno == "Tarde":
                celda.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            elif turno == "Part-Time":
                celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            elif turno == "Libre":
                celda.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        fila += 1
    
    # Ajustar anchos
    ws_horarios.column_dimensions['A'].width = 20
    for i in range(2, 9):
        col = get_column_letter(i)
        ws_horarios.column_dimensions[col].width = 14
    
    # === HOJA 3: Leyenda ===
    ws_leyenda = wb.create_sheet("Leyenda")
    
    ws_leyenda['A1'] = "LEYENDA DE COLORES"
    ws_leyenda['A1'].font = titulo_font
    ws_leyenda['A1'].fill = titulo_fill
    ws_leyenda.merge_cells('A1:B1')
    
    leyenda = [
        ("Mañana", "FFF2CC"),
        ("Intermedio", "E2EFDA"),
        ("Tarde", "FCE4D6"),
        ("Part-Time", "DDEBF7"),
        ("Libre", "F2F2F2"),
        ("-", "FFFFFF")
    ]
    
    fila = 3
    for turno, color in leyenda:
        ws_leyenda[f'A{fila}'] = turno
        ws_leyenda[f'A{fila}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws_leyenda[f'A{fila}'].border = border
        ws_leyenda[f'A{fila}'].alignment = center_align
        
        if turno == "-":
            ws_leyenda[f'B{fila}'] = "No trabaja (Part-Time días de semana)"
        elif turno == "Libre":
            ws_leyenda[f'B{fila}'] = "Día de descanso"
        else:
            ws_leyenda[f'B{fila}'] = f"Turno de {turno.lower()}"
        
        ws_leyenda[f'B{fila}'].border = border
        fila += 1
    
    ws_leyenda.column_dimensions['A'].width = 15
    ws_leyenda.column_dimensions['B'].width = 35
    
    # Guardar archivo
    wb.save(ruta_archivo)
